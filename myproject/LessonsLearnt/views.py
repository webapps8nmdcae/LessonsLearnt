import requests
from flask import Blueprint, render_template, jsonify, request, current_app
from openpyxl.styles import PatternFill

from myproject.models import *
import os
import json
import csv
from io import StringIO, BytesIO
from datetime import datetime
from flask import Blueprint, render_template, jsonify, request, current_app, Response, send_file, session, request, redirect
from myproject import base_url
import requests
lessons_learnt_blueprint = Blueprint('LessonsLearnt', __name__, template_folder='../templates')


@lessons_learnt_blueprint.route('/login')
def login():
    token = request.args.get('token')
    session['token'] = token # to get the token --> tokentest = session['token'] ////  print(tokentest)

    # 2. Get the app id from the URL & check access
    appid = request.args.get('appid')
    url1 = base_url + '/DataAPI/api/Apps/GetUserAppAcces?AppId=' + appid
    payload = {}
    headers = {'Authorization': 'Bearer ' + token}
    response1 = requests.request("GET", url1, headers=headers, data=payload)

    if response1.text is not '':  # access
        # session['loggings_token'] = token
        user = User.query.filter_by(id=response1.text).first()
      
        # 3. Login the user
        session.permanent = True
        session['lessonslearnt_app'] = user.serialize
        session['base_url'] = base_url
        session['ApiUrl'] = base_url + "/DataAPI/api/"
        return redirect('/lessonslearnt/newlesson')
    else:
        session['access'] = 'false'
        notification = "Access to the application was rejected. Please contact dmwebapps@nmdc-group.com for assistance"
        return redirect(base_url + '?noti='+notification)

@lessons_learnt_blueprint.route('/home')
def home():
    if 'lessonslearnt_app' not in session:
        return redirect(base_url)
    return render_template('LessonsLearnt/Home.html')

@lessons_learnt_blueprint.route('/newlesson')
def newlesson():
    if 'lessonslearnt_app' not in session:
        return redirect(base_url)
    return render_template('LessonsLearnt/newlesson.html')

@lessons_learnt_blueprint.route('/viewlesson')
def viewlesson():
    if 'lessonslearnt_app' not in session:
        return redirect(base_url)
    return render_template('LessonsLearnt/viewlesson.html')

@lessons_learnt_blueprint.route('/fuel_management')
def fuel_management():
    if 'lessonslearnt_app' not in session:
        return redirect(base_url)
    return render_template('LessonsLearnt/fuel_management.html')

@lessons_learnt_blueprint.route('/GetMasterData', methods=['GET'])
def GetMasterData():
    
    user_id = session.get('lessonslearnt_app', {}).get('id')

    if not user_id:
        return jsonify({'error': 'Unauthorized'}), 401
        
    projects = db.session.query(Project).filter(Project.Relevance_LessonsLearnt == 1).order_by(Project.ProjectNumber, Project.ProjectName).all()
    clients = db.session.query(ProjectEmployerClassification).order_by(ProjectEmployerClassification.ProjectEmployerClassificationName).all()
    categories = db.session.query(LessonLearnCategory, LLCategoryType).join(LLCategoryType, LessonLearnCategory.qhse_LLCategoryTypeId == LLCategoryType.Id).order_by(LessonLearnCategory.qhse_LLCategoryTypeId, LessonLearnCategory.Category).all()
    subcategories = db.session.query(LessonLearnSubCategory, LessonLearnCategory, LLCategoryType).join(LessonLearnCategory, LessonLearnSubCategory.qhse_LessonLearnCategoryId == LessonLearnCategory.Id).join(LLCategoryType, LessonLearnCategory.qhse_LLCategoryTypeId == LLCategoryType.Id).order_by(LessonLearnSubCategory.qhse_LessonLearnCategoryId, LessonLearnSubCategory.SubCategory).all()
    priorities = db.session.query(LessonLearnPriorityLevel).order_by(LessonLearnPriorityLevel.Id).all()
    impactareas = db.session.query(LessonLearnPrimaryBusinessImpact).order_by(LessonLearnPrimaryBusinessImpact.Id).all()
    
    users = db.session.query(User).all()
    user_dict = {str(u.id): u.FullName for u in users if u.FullName}

    def get_expertise_names(expertise_id_str):
        if not expertise_id_str:
            return ""
        ids = [i.strip() for i in expertise_id_str.split(',')]
        names = [user_dict.get(i, '') for i in ids]
        return ", ".join([n for n in names if n])
    
    project_list = [{'Id': p.Id, 'Name': p.ProjectName, 'ProjectCode': p.ProjectNumber, 'ClassificationId': p.proj_ProjectEmployerClassificationId} for p in projects]
    client_list = [{'Id': c.Id, 'Name': c.ProjectEmployerClassificationName} for c in clients]
    category_list = [{'Id': c.LessonLearnCategory.Id, 'Category': c.LessonLearnCategory.Category, 'CategoryTypeId': c.LessonLearnCategory.qhse_LLCategoryTypeId, 'CategoryType': c.LLCategoryType.CategoryType} for c in categories]
    subcategory_list = [{'Id': sc.LessonLearnSubCategory.Id, 'SubCategory': sc.LessonLearnSubCategory.SubCategory, 'CategoryId': sc.LessonLearnSubCategory.qhse_LessonLearnCategoryId, 'Category': sc.LessonLearnCategory.Category, 'CategoryTypeId': sc.LessonLearnCategory.qhse_LLCategoryTypeId, 'CategoryType': sc.LLCategoryType.CategoryType, 'ExpertiseId': sc.LessonLearnSubCategory.ExpertiseId, 'ExpertiseName': get_expertise_names(sc.LessonLearnSubCategory.ExpertiseId)} for sc in subcategories]
    priorities_list = [{'Id': p.Id, 'PriorityLevel': p.PriorityLevel} for p in priorities]
    impactareas_list = [{'Id': i.Id, 'PrimaryBusinessImpact': i.PrimaryBusinessImpact} for i in impactareas]
    
    return jsonify({
        'projects': project_list,
        'clients': client_list,
        'categories': category_list,
        'subcategories': subcategory_list,
        'priorities': priorities_list,
        'impactareas': impactareas_list
    }), 200

@lessons_learnt_blueprint.route('/SaveLessonLearn', methods=['POST'])
def SaveLessonLearn():    
    try:
        createdbyid = session.get('lessonslearnt_app', {}).get('id')
        origin_type = request.form.get('originType')
        project_id = request.form.get('project')
        client_id = request.form.get('client')
        category_id = request.form.get('category')
        background = request.form.get('background')
        lesson_learnt = request.form.get('lessonLearnt')
        recommendations = request.form.get('recommendations')
        priority_id = request.form.get('priority')
        impact_area_id = request.form.get('impact')
        
        form_category_data_str = request.form.get('formCategoryData')
        form_category_data = []
        if form_category_data_str:
            form_category_data = json.loads(form_category_data_str)
            
        expertise_ids_str = request.form.get('expertiseIds')
        expertise_ids = []
        if expertise_ids_str:
            expertise_ids = [i.strip() for i in expertise_ids_str.split(',') if i.strip()]
            
        active_tab_id = request.form.get('activeTabId')
        
        is_project = 1 if origin_type == 'project' else 0
        now = datetime.now()
        
        new_form = LessonLearnForm(
            proj_ProjectId=project_id if project_id else None,
            Date=now,
            qhse_LLCategoryId=category_id if category_id else None,
            Background=background,
            LessonLearnt=lesson_learnt,
            qshe_LLApprovalOrderId=0,
            ActionStatus=0,
            RequesterId=5864,
            IsProject=bool(is_project),
            proj_ProjectEmployerClassificationId=client_id if client_id else None,
            IsArchived=0,
            CreatedBy=createdbyid,
            CreatedOn=now
        )
        db.session.add(new_form)
        db.session.flush() # to get new_form.Id
        form_id = new_form.Id
        
        for item in form_category_data:
            new_cat = LessonLearnFormCategory(
                qhse_LessonLearnFormId=form_id,
                qhse_LLCategoryTypeId=item.get('categoryTypeId'),
                qhse_LLCategoryId=item.get('categoryId'),
                qhse_LLSubCategoryId=item.get('subCategoryId'),
                CreatedBy=createdbyid,
                CreatedOn=now
            )
            db.session.add(new_cat)
            
        for exp_id in expertise_ids:
            new_exp = LessonLearnFormExpertise(
                ExpertiseId=exp_id,
                qhse_LessonLearnFormId=form_id,
                CreatedBy=createdbyid,
                CreatedOn=now
            )
            db.session.add(new_exp)
        
        documents = request.files.getlist('document')
        static_folder = current_app.static_folder
        if not static_folder:
            static_folder = os.path.join(current_app.root_path, 'static')
            
        upload_folder = os.path.join(static_folder, 'LessonsLearntAttachments')
        os.makedirs(upload_folder, exist_ok=True)
        
        seq = 1
        for file in documents:
            if file and file.filename:
                ext = os.path.splitext(file.filename)[1]
                filename = f"{form_id}_{now.strftime('%Y%m%d%H%M%S')}_{seq}{ext}"
                file_path = os.path.join(upload_folder, filename)
                file.save(file_path)
                
                db_path = f"LessonsLearntAttachments/{filename}"
                
                new_doc = LessonLearnFormDocuments(
                    qhse_LLFormId=form_id,
                    AttachmentPath=db_path
                )
                db.session.add(new_doc)
                seq += 1
                
        new_approval = LessonLearnLApprovalData(
            LLFormId=form_id,
            qhse_LLApprovalOrderId=0,
            qhse_LLDepartmentId=0,
            hr_EmployeeId=5864,
            Comments=recommendations,
            qshe_LLApprovalStatusId=0,
            CreatedBy=createdbyid,
            CreatedOn=now,
            qhse_LessonLearnPriorityLevelId = priority_id,
            qhse_LessonLearnPrimaryBusinessImpactId = impact_area_id
        )
        db.session.add(new_approval)
        db.session.commit()
        return jsonify({'status': 'success', 'message': 'Lesson learnt saved successfully.', 'id': form_id}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@lessons_learnt_blueprint.route('/SearchLessonLearn', methods=['POST'])
def SearchLessonLearn():
    try:
        data = request.json
        client_id = data.get('client')
        project_id = data.get('project')
        category_id = data.get('category')
        subcategory_id = data.get('subCategory')
        status = data.get('status')
        
        query = db.session.query(
            LessonLearnForm, 
            Project.ProjectNumber,
            Project.ProjectName,
            ProjectEmployerClassification.ProjectEmployerClassificationName,
            LessonLearnLApprovalData.Comments,
            LessonLearnPriorityLevel.PriorityLevel,
            LessonLearnPrimaryBusinessImpact.PrimaryBusinessImpact
        ).outerjoin(Project, LessonLearnForm.proj_ProjectId == Project.Id)\
         .outerjoin(ProjectEmployerClassification, LessonLearnForm.proj_ProjectEmployerClassificationId == ProjectEmployerClassification.Id)\
         .outerjoin(LessonLearnLApprovalData, LessonLearnForm.Id == LessonLearnLApprovalData.LLFormId)\
         .outerjoin(LessonLearnPriorityLevel, LessonLearnLApprovalData.qhse_LessonLearnPriorityLevelId == LessonLearnPriorityLevel.Id)\
         .outerjoin(LessonLearnPrimaryBusinessImpact, LessonLearnLApprovalData.qhse_LessonLearnPrimaryBusinessImpactId == LessonLearnPrimaryBusinessImpact.Id)
        
        if client_id:
            query = query.filter(LessonLearnForm.proj_ProjectEmployerClassificationId == client_id)
        if project_id:
            query = query.filter(LessonLearnForm.proj_ProjectId == project_id)
        if status:
            if status == 'pending':
                query = query.filter(LessonLearnForm.ActionStatus == 0)
            elif status == 'completed':
                query = query.filter(LessonLearnForm.ActionStatus == 1)
            elif status == 'rejected':
                query = query.filter(LessonLearnForm.ActionStatus == 2)

        if category_id or subcategory_id:
            query = query.join(LessonLearnFormCategory, LessonLearnForm.Id == LessonLearnFormCategory.qhse_LessonLearnFormId)
            if category_id:
                query = query.filter(LessonLearnFormCategory.qhse_LLCategoryId == category_id)
            if subcategory_id:
                query = query.filter(LessonLearnFormCategory.qhse_LLSubCategoryId == subcategory_id)
                
        results = query.order_by(LessonLearnForm.Id.desc(), LessonLearnForm.CreatedOn.desc()).all()
        
        if not results:
            return jsonify({'status': 'info', 'message': 'No data found.'}), 200
        
        users = db.session.query(User).all()
        user_dict = {str(u.id): u.FullName for u in users if u.FullName}

        formatted_results = []
        for row in results:
            form = row.LessonLearnForm
            
            # Fetch all categories for this form to display
            form_categories = db.session.query(
                LLCategoryType.CategoryType,
                LessonLearnCategory.Category,
                LessonLearnSubCategory.SubCategory
            ).select_from(LessonLearnFormCategory)\
             .join(LessonLearnCategory, LessonLearnFormCategory.qhse_LLCategoryId == LessonLearnCategory.Id)\
             .join(LLCategoryType, LessonLearnFormCategory.qhse_LLCategoryTypeId == LLCategoryType.Id)\
             .outerjoin(LessonLearnSubCategory, LessonLearnFormCategory.qhse_LLSubCategoryId == LessonLearnSubCategory.Id)\
             .filter(LessonLearnFormCategory.qhse_LessonLearnFormId == form.Id).all()
             
            category_types = "<br>".join(list(set([c.CategoryType for c in form_categories if c.CategoryType])))
            
            cat_subcats = []
            seen = set()
            for c in form_categories:
                identifier = f"{c.Category}|{c.SubCategory}"
                if identifier not in seen:
                    seen.add(identifier)
                    cat_subcats.append({
                        'Category': c.Category,
                        'SubCategory': c.SubCategory if c.SubCategory else ''
                    })
            
            # Fetch attachments
            docs = db.session.query(LessonLearnFormDocuments).filter(LessonLearnFormDocuments.qhse_LLFormId == form.Id).all()
            attachments = [{'Id': d.Id, 'Path': d.AttachmentPath} for d in docs]
            
            status_text = 'Pending Approval'
            if form.ActionStatus == 1:
                status_text = 'Completed'
            elif form.ActionStatus == 2:
                status_text = 'Rejected'
                
            form_expertise_records = db.session.query(LessonLearnFormExpertise).filter(LessonLearnFormExpertise.qhse_LessonLearnFormId == form.Id).all()
            expertise_names = [user_dict.get(str(exp.ExpertiseId), str(exp.ExpertiseId)) for exp in form_expertise_records if exp.ExpertiseId]
            expertise_text = ", ".join([n for n in expertise_names if n]) if expertise_names else '-'

            formatted_results.append({
                'Id': form.Id,
                'LessonCategory': 'Project' if form.IsProject else 'Department',
                'Project': f"{row.ProjectNumber} - {row.ProjectName}" if row.ProjectName else '-',
                'Client': row.ProjectEmployerClassificationName if row.ProjectEmployerClassificationName else '-',
                'Background': form.Background,
                'LessonLearnt': form.LessonLearnt,
                'CategoryType': category_types,
                'CategorySubCategory': cat_subcats,
                'PrimaryBusinessImpact': row.PrimaryBusinessImpact if row.PrimaryBusinessImpact else '-',
                'PriorityLevel': row.PriorityLevel if row.PriorityLevel else '-',
                'Expertise': expertise_text,
                'Recommendations': row.Comments if row.Comments else '-',
                'Status': status_text,
                'Attachments': attachments,
                'CreatedOn': form.CreatedOn.strftime('%d-%b-%Y %I:%M %p') if form.CreatedOn else '-',
                'CreatedBy': user_dict.get(str(form.CreatedBy), form.CreatedBy) if form.CreatedBy else '-',
                'UpdatedOn': form.UpdatedOn.strftime('%d-%b-%Y %I:%M %p') if form.UpdatedOn else '-',
                'UpdatedBy': user_dict.get(str(form.UpdatedBy), form.UpdatedBy) if form.UpdatedBy else '-',
                'LessonLearnedDate': form.Date.strftime('%d-%b-%Y %I:%M %p') if form.Date else '-'
            })
            
        return jsonify({'status': 'success', 'data': formatted_results}), 200

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@lessons_learnt_blueprint.route('/DeleteLessonLearn', methods=['POST'])
def DeleteLessonLearn():
    try:
        data = request.json
        lesson_id = data.get('id')
        if not lesson_id:
            return jsonify({'status': 'error', 'message': 'Invalid ID.'}), 400
            
        form = db.session.query(LessonLearnForm).filter(LessonLearnForm.Id == lesson_id).first()
        if not form:
            return jsonify({'status': 'error', 'message': 'Record not found.'}), 404

        # Delete from LessonLearnFormCategory
        db.session.query(LessonLearnFormCategory).filter(LessonLearnFormCategory.qhse_LessonLearnFormId == lesson_id).delete()
        
        # Delete from LessonLearnLApprovalData
        db.session.query(LessonLearnLApprovalData).filter(LessonLearnLApprovalData.LLFormId == lesson_id).delete()
        
        # Delete files from static folder and LessonLearnFormDocuments
        documents = db.session.query(LessonLearnFormDocuments).filter(LessonLearnFormDocuments.qhse_LLFormId == lesson_id).all()
        static_folder = current_app.static_folder
        if not static_folder:
            static_folder = os.path.join(current_app.root_path, 'static')
            
        for doc in documents:
            file_path = os.path.join(static_folder, doc.AttachmentPath)
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except OSError:
                    pass # Ignore file not found or permission errors on deletion
            db.session.delete(doc)
            
        # Delete from LessonLearnForm
        db.session.delete(form)
        db.session.commit()
        
        return jsonify({'status': 'success', 'message': 'Deleted successfully.'}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@lessons_learnt_blueprint.route('/ExportLessonsLearnt', methods=['POST'])
def ExportLessonsLearnt():
    try:
        data = request.json
        client_id = data.get('client')
        project_id = data.get('project')
        category_id = data.get('category')
        subcategory_id = data.get('subCategory')
        status = data.get('status')
        
        query = db.session.query(
            LessonLearnForm, 
            Project.ProjectName,
            ProjectEmployerClassification.ProjectEmployerClassificationName,
            LessonLearnLApprovalData.Comments,
            LessonLearnPriorityLevel.PriorityLevel,
            LessonLearnPrimaryBusinessImpact.PrimaryBusinessImpact
        ).outerjoin(Project, LessonLearnForm.proj_ProjectId == Project.Id)\
         .outerjoin(ProjectEmployerClassification, LessonLearnForm.proj_ProjectEmployerClassificationId == ProjectEmployerClassification.Id)\
         .outerjoin(LessonLearnLApprovalData, LessonLearnForm.Id == LessonLearnLApprovalData.LLFormId)\
         .outerjoin(LessonLearnPriorityLevel, LessonLearnLApprovalData.qhse_LessonLearnPriorityLevelId == LessonLearnPriorityLevel.Id)\
         .outerjoin(LessonLearnPrimaryBusinessImpact, LessonLearnLApprovalData.qhse_LessonLearnPrimaryBusinessImpactId == LessonLearnPrimaryBusinessImpact.Id)
        
        if client_id:
            query = query.filter(LessonLearnForm.proj_ProjectEmployerClassificationId == client_id)
        if project_id:
            query = query.filter(LessonLearnForm.proj_ProjectId == project_id)
        if status:
            if status == 'pending':
                query = query.filter(LessonLearnForm.ActionStatus == 0)
            elif status == 'completed':
                query = query.filter(LessonLearnForm.ActionStatus == 1)
            elif status == 'rejected':
                query = query.filter(LessonLearnForm.ActionStatus == 2)
            
        if category_id or subcategory_id:
            query = query.join(LessonLearnFormCategory, LessonLearnForm.Id == LessonLearnFormCategory.qhse_LessonLearnFormId)
            if category_id:
                query = query.filter(LessonLearnFormCategory.qhse_LLCategoryId == category_id)
            if subcategory_id:
                query = query.filter(LessonLearnFormCategory.qhse_LLSubCategoryId == subcategory_id)
                
        results = query.order_by(LessonLearnForm.Id.desc(), LessonLearnForm.CreatedOn.desc()).all()
        
        if not results:
            return jsonify({'status': 'info', 'message': 'No data found for the selected criteria.'}), 200
        
        users = db.session.query(User).all()
        user_dict = {str(u.id): u.FullName for u in users if u.FullName}

        # We will use openpyxl for Excel formatting
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return jsonify({'status': 'error', 'message': 'openpyxl library is not installed. Cannot generate Excel.'}), 500

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lessons Learnt"
        
        headers = [
            'ID', 'Lesson Category', 'Project', 'Client', 'Background Context', 'Lesson Learnt', 
            'Category Type', 'Category - Sub Category', 'Status', 'Created On', 'Created By', 'Updated On', 'Updated By', 'Date'
        ]
        ws.append(headers)
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Format headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            # Light background color
            cell.fill = PatternFill(
                start_color="4F81BD",  # Steel blue (Excel-style)
                end_color="4F81BD",
                fill_type="solid"
            )
        
        for row_idx, row in enumerate(results, 2):
            form = row.LessonLearnForm
            
            form_categories = db.session.query(
                LLCategoryType.CategoryType,
                LessonLearnCategory.Category,
                LessonLearnSubCategory.SubCategory
            ).select_from(LessonLearnFormCategory)\
             .join(LessonLearnCategory, LessonLearnFormCategory.qhse_LLCategoryId == LessonLearnCategory.Id)\
             .join(LLCategoryType, LessonLearnFormCategory.qhse_LLCategoryTypeId == LLCategoryType.Id)\
             .outerjoin(LessonLearnSubCategory, LessonLearnFormCategory.qhse_LLSubCategoryId == LessonLearnSubCategory.Id)\
             .filter(LessonLearnFormCategory.qhse_LessonLearnFormId == form.Id).all()
             
            category_types = ", ".join(list(set([c.CategoryType for c in form_categories if c.CategoryType])))
            
            cat_subcats = []
            seen = set()
            for c in form_categories:
                identifier = f"{c.Category} - {c.SubCategory}" if c.SubCategory else c.Category
                if identifier not in seen:
                    seen.add(identifier)
                    cat_subcats.append(identifier)
            category_subcategories = " | ".join(cat_subcats)
            
            status_text = 'Pending Approval'
            if form.ActionStatus == 1:
                status_text = 'Completed'
            elif form.ActionStatus == 2:
                status_text = 'Rejected'
                
            row_data = [
                form.Id,
                'Project' if form.IsProject else 'Department',
                row.ProjectName if row.ProjectName else '-',
                row.ProjectEmployerClassificationName if row.ProjectEmployerClassificationName else '-',
                form.Background,
                form.LessonLearnt,
                category_types,
                category_subcategories,
                status_text,
                form.CreatedOn if form.CreatedOn else None,
                user_dict.get(str(form.CreatedBy), form.CreatedBy) if form.CreatedBy else '-',
                form.UpdatedOn if form.UpdatedOn else None,
                user_dict.get(str(form.UpdatedBy), form.UpdatedBy) if form.UpdatedBy else '-',
                form.Date if form.Date else None
            ]
            
            date_columns = [10, 12, 14]  # Created On, Updated On, Lesson Learned Date
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                if col_idx in date_columns and value:
                    cell.number_format = 'dd-mmm-yy h:mm AM/PM'

        # Auto expand columns
        for col_num, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            for cell in column_cells:
                if cell.value:
                    lines = str(cell.value).split('\n')
                    for line in lines:
                        if len(line) > max_length:
                            max_length = len(line)
            # Cap maximum width to 50 so very long text doesn't make columns too wide
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_num)].width = adjusted_width
            
        # Add auto filter
        ws.auto_filter.ref = ws.dimensions
            
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        
        return send_file(
            out,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='LessonsLearnt_Export.xlsx'
        )

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500